"""2024 workbook adapter for non-agricultural CU baseline inputs.

This is deliberately the only module that knows legacy worksheet cells.  It
reads displayed input values, not Excel formula text; production calculation
is performed by :mod:`non_agricultural_use`.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from .non_agricultural_use import (
    GILA_EXCLUSIVE_OF_VIRDEN,
    SAN_FRANCISCO,
    SAN_SIMON,
    VIRDEN,
    CliffGilaMunicipalUse,
    FreeportAccounting,
    LakeEvaporationSite,
    LivestockInventory,
    MunicipalDiversion,
    StockTankSite,
)


@dataclass(frozen=True)
class LegacyNonAgriculturalInputs:
    livestock: LivestockInventory
    stock_tanks: tuple[StockTankSite, ...]
    lakes: tuple[LakeEvaporationSite, ...]
    municipal_diversions: tuple[MunicipalDiversion, ...]
    cliff_gila_municipal: CliffGilaMunicipalUse


def read_2024_non_agricultural_inputs(path: str | Path) -> LegacyNonAgriculturalInputs:
    """Read all raw/calibrated 2024 non-agricultural source inputs."""

    try:
        import openpyxl
    except ImportError as error:  # pragma: no cover - project dependency
        raise RuntimeError("Reading legacy non-agricultural inputs requires openpyxl.") from error

    workbook = openpyxl.load_workbook(Path(path), data_only=True, read_only=False)
    sheet = workbook["stock-dom-com-ind-evap"]
    freeport = workbook["Freeport"]
    cliff_gila = workbook["cliffgila"]
    return LegacyNonAgriculturalInputs(
        livestock=LivestockInventory(
            cattle_by_county={"CATRON": _number(sheet["B4"].value), "GRANT": _number(sheet["B5"].value), "HIDALGO": _number(sheet["B6"].value)},
            sheep_by_county={"CATRON": _number(sheet["C4"].value), "GRANT": _zero(sheet["C5"].value), "HIDALGO": _zero(sheet["C6"].value)},
            national_forest_cattle_head_months=_number(sheet["E10"].value),
        ),
        stock_tanks=(
            _stock_tank(sheet, 115, "Luna", SAN_FRANCISCO),
            _stock_tank(sheet, 116, "Apache-Aragon", SAN_FRANCISCO),
            _stock_tank(sheet, 117, "Reserve", SAN_FRANCISCO),
            _stock_tank(sheet, 118, "Glenwood", SAN_FRANCISCO),
            _stock_tank(sheet, 123, "Gila Hot Springs (San Francisco)", SAN_FRANCISCO),
            _stock_tank(sheet, 130, "Upper Gila", GILA_EXCLUSIVE_OF_VIRDEN),
            _stock_tank(sheet, 131, "Cliff-Gila-Redrock", GILA_EXCLUSIVE_OF_VIRDEN),
            _stock_tank(sheet, 136, "Gila Hot Springs (Gila)", GILA_EXCLUSIVE_OF_VIRDEN),
            _stock_tank(sheet, 142, "Virden", VIRDEN),
            _stock_tank(sheet, 146, "San Simon", SAN_SIMON),
        ),
        lakes=(
            LakeEvaporationSite("Lake Roberts", GILA_EXCLUSIVE_OF_VIRDEN, _number(sheet["B81"].value), _number(sheet["C81"].value), _number(sheet["E81"].value)),
            LakeEvaporationSite("Wall and Snow", GILA_EXCLUSIVE_OF_VIRDEN, _number(sheet["B82"].value), _number(sheet["C82"].value), _number(sheet["E82"].value)),
            LakeEvaporationSite("Bill Evans (NM Game and Fish allocation)", GILA_EXCLUSIVE_OF_VIRDEN, fixed_use_af=_number(sheet["F86"].value)),
        ),
        municipal_diversions=(
            MunicipalDiversion("Luna", SAN_FRANCISCO, _number(sheet["C57"].value)),
            MunicipalDiversion("Apache-Aragon", SAN_FRANCISCO, _number(sheet["C58"].value)),
            MunicipalDiversion("Reserve", SAN_FRANCISCO, _number(sheet["C59"].value)),
            MunicipalDiversion("Glenwood", SAN_FRANCISCO, _number(sheet["C60"].value)),
            MunicipalDiversion("Upper Gila geothermal", GILA_EXCLUSIVE_OF_VIRDEN, _number(sheet["C63"].value), 82.59),
            MunicipalDiversion("Redrock", GILA_EXCLUSIVE_OF_VIRDEN, _number(sheet["C65"].value)),
            MunicipalDiversion("San Simon", SAN_SIMON, _number(sheet["C68"].value)),
            MunicipalDiversion("Virden", VIRDEN, _number(sheet["C70"].value)),
        ),
        cliff_gila_municipal=CliffGilaMunicipalUse(
            fish_pond_diversion_af=73.16,
            fish_pond_nonconsumptive_af=0.23,
            exported_to_mimbres_af=663.33,
            freeport=FreeportAccounting(
                tyrone_wells_af=_number(freeport["B12"].value),
                evans_reservoir_to_mine_af=_number(freeport["B14"].value),
                bill_evans_evaporation_af=_number(freeport["B15"].value),
                t_irrigation_cu_af=_number(freeport["B17"].value),
                t13_usfs_diversion_af=_number(freeport["B18"].value),
                seepage_credit_af=_number(freeport["B19"].value),
            ),
            fish_pond_evaporation_allocation_af=_number(cliff_gila["W24"].value),
        ),
    )


def _stock_tank(sheet: object, row: int, name: str, stream_system: str) -> StockTankSite:
    return StockTankSite(
        name=name,
        stream_system=stream_system,
        adjusted_pan_evaporation_in=_number(sheet.cell(row, 2).value),
        precipitation_in=_number(sheet.cell(row, 3).value),
        average_surface_area_acres=_number(sheet.cell(row, 5).value),
        tank_count=_number(sheet.cell(row, 7).value),
    )


def _number(value: object) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        return float(value)
    raise ValueError(f"Expected numeric legacy value, got {value!r}")


def _zero(value: object) -> float:
    return 0.0 if value is None else _number(value)

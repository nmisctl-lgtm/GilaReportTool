"""Extract auditable 2024 metered-ditch assets from the report workbook.

This is a baseline reader, not the future report generator.  It exposes the
crop/reservoir/pre-plant acreage and the existing shortage-cell treatment so
those manual policy decisions can be moved into a reviewable input table.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from .diversion_mapping import SourceDitchMapping
from .diversion_ledger import DitchInput, QAIssue


AREA_SHEET_NAMES = {
    "LUNA": "luna",
    "APACHE-ARAGON": "apachearagon",
    "RESERVE": "reserve",
    "GLENWOOD": "glenwood",
    "UPPER GILA": "uppergila",
    "CLIFF-GILA": "cliffgila",
    "REDROCK": "redrock",
    "VIRDEN VALLEY": "virden",
    "SAN SIMON": "sansimon",
}


@dataclass(frozen=True)
class LegacyMeteredDitchAsset:
    """One metered-ditch block as stored in the 2024 report workbook."""

    canonical_ditch_id: str
    area_name: str
    report_ditch_name: str
    crop_acres: float
    reservoir_acres: float
    preplant_acres: float
    monthly_diversion_acft: tuple[float, ...]
    monthly_reservoir_net_evap_acft: tuple[float, ...]
    reservoir_net_evap_formula_months: tuple[bool, ...]
    monthly_diversion_required_acft: tuple[float, ...]
    requirement_formula_months: tuple[bool, ...]
    shortage_formula_months: tuple[bool, ...]
    monthly_shortage_acft: tuple[float, ...]


@dataclass(frozen=True)
class LegacyAreaDiversionInputs:
    """The report-sheet inputs shared by all metered ditches in one area."""

    area_name: str
    efficiency: float
    monthly_report_cir_ft: tuple[float, ...]
    monthly_adjusted_pan_evap_ft: tuple[float, ...]
    monthly_precip_ft: tuple[float, ...]


def read_2024_metered_ditch_assets(
    path: str | Path, mappings: Iterable[SourceDitchMapping]
) -> tuple[LegacyMeteredDitchAsset, ...]:
    """Read configured 2024 report blocks without inferring policy semantics."""

    try:
        import openpyxl
    except ImportError as error:  # pragma: no cover - project dependency
        raise RuntimeError("Reading legacy report assets requires openpyxl.") from error

    mappings = tuple(mapping for mapping in mappings if mapping.disposition == "report_ditch")
    values = openpyxl.load_workbook(Path(path), data_only=True, read_only=False)
    formulas = openpyxl.load_workbook(Path(path), data_only=False, read_only=False)
    assets: list[LegacyMeteredDitchAsset] = []
    for mapping in mappings:
        sheet_name = AREA_SHEET_NAMES[mapping.area_name or ""]
        value_sheet = values[sheet_name]
        formula_sheet = formulas[sheet_name]
        diversion_column = _find_report_ditch_column(value_sheet, mapping.report_ditch_name or "")
        acre_column = diversion_column + 1
        monthly_diversion = tuple(_number(value_sheet.cell(row, diversion_column).value) for row in range(12, 24))
        reservoir_net_evap = tuple(_number(value_sheet.cell(row, diversion_column + 1).value) for row in range(12, 24))
        monthly_required = tuple(_number(value_sheet.cell(row, diversion_column + 4).value) for row in range(12, 24))
        monthly_shortage = tuple(_number(value_sheet.cell(row, diversion_column + 5).value) for row in range(12, 24))
        requirement_formula_months = tuple(
            isinstance(formula_sheet.cell(row, diversion_column + 4).value, str)
            and formula_sheet.cell(row, diversion_column + 4).value.startswith("=")
            for row in range(12, 24)
        )
        reservoir_net_evap_formula_months = tuple(
            isinstance(formula_sheet.cell(row, diversion_column + 1).value, str)
            and formula_sheet.cell(row, diversion_column + 1).value.startswith("=")
            for row in range(12, 24)
        )
        shortage_formula_months = tuple(
            isinstance(formula_sheet.cell(row, diversion_column + 5).value, str)
            and formula_sheet.cell(row, diversion_column + 5).value.startswith("=")
            for row in range(12, 24)
        )
        assets.append(LegacyMeteredDitchAsset(
            canonical_ditch_id=mapping.canonical_ditch_id or "",
            area_name=mapping.area_name or "",
            report_ditch_name=mapping.report_ditch_name or "",
            reservoir_acres=_number(value_sheet.cell(7, acre_column).value),
            preplant_acres=_number(value_sheet.cell(8, acre_column).value),
            crop_acres=_number(value_sheet.cell(9, acre_column).value),
            monthly_diversion_acft=monthly_diversion,
            monthly_reservoir_net_evap_acft=reservoir_net_evap,
            reservoir_net_evap_formula_months=reservoir_net_evap_formula_months,
            monthly_diversion_required_acft=monthly_required,
            requirement_formula_months=requirement_formula_months,
            shortage_formula_months=shortage_formula_months,
            monthly_shortage_acft=monthly_shortage,
        ))
    return tuple(assets)


def read_2024_area_diversion_inputs(path: str | Path) -> tuple[LegacyAreaDiversionInputs, ...]:
    """Read the common efficiency, adjusted pan evaporation, and precipitation.

    The legacy sheets use a pan-evaporation coefficient of 0.8 before the
    reservoir balance.  The extracted pan values are therefore already
    adjusted and can be used directly in ``calculate_area_diversion_ledger``.
    """

    try:
        import openpyxl
    except ImportError as error:  # pragma: no cover - project dependency
        raise RuntimeError("Reading legacy report assets requires openpyxl.") from error

    workbook = openpyxl.load_workbook(Path(path), data_only=True, read_only=False)
    areas: list[LegacyAreaDiversionInputs] = []
    for area_name, sheet_name in AREA_SHEET_NAMES.items():
        sheet = workbook[sheet_name]
        efficiency = _number(sheet["D3"].value)
        areas.append(LegacyAreaDiversionInputs(
            area_name=area_name,
            efficiency=efficiency,
            # The legacy column B is explicitly labeled "CIR (ft)/e".
            # Convert it back to CIR before passing it to the ledger, which
            # itself performs the division by efficiency.
            monthly_report_cir_ft=tuple(
                _number(sheet.cell(row, 2).value) * efficiency for row in range(12, 24)
            ),
            monthly_adjusted_pan_evap_ft=tuple(_number(sheet.cell(row, 3).value) for row in range(12, 24)),
            monthly_precip_ft=tuple(_number(sheet.cell(row, 4).value) for row in range(12, 24)),
        ))
    return tuple(areas)


def build_historical_ditch_inputs(
    assets: Iterable[LegacyMeteredDitchAsset],
) -> tuple[DitchInput, ...]:
    """Turn 2024 evidence into ledger inputs for workbook-parity validation.

    All blocks represented here are sourced from the legacy daily metered-flow
    table.  The historical formula/non-formula pattern becomes the explicit
    2024 shortage-assessment switch; it is not a policy inference for 2025.
    """

    return tuple(DitchInput(
        ditch_id=asset.canonical_ditch_id,
        name=asset.report_ditch_name,
        crop_acres=asset.crop_acres,
        reservoir_acres=asset.reservoir_acres,
        monthly_diversion_acft=asset.monthly_diversion_acft,
        measurement_status=("metered",) * 12,
        shortage_assessed=asset.shortage_formula_months,
        monthly_reservoir_net_evap_override_acft=tuple(
            None if is_formula else value
            for is_formula, value in zip(
                asset.reservoir_net_evap_formula_months, asset.monthly_reservoir_net_evap_acft
            )
        ),
    ) for asset in assets)


def validate_historical_requirement_overrides(
    assets: Iterable[LegacyMeteredDitchAsset],
) -> tuple[QAIssue, ...]:
    """Flag nonzero requirement cells that cannot be reproduced from formulas.

    A historical zero without a formula can be an intentional no-shortage
    policy.  A nonzero requirement constant, however, changes the demand
    calculation itself and must be moved into a named, approved exception
    before an automated report can use it.
    """

    issues: list[QAIssue] = []
    for asset in assets:
        for month, (is_formula, required) in enumerate(
            zip(asset.requirement_formula_months, asset.monthly_diversion_required_acft), 1
        ):
            if not is_formula and required != 0:
                issues.append(QAIssue(
                    "error", "manual_requirement_override",
                    "Nonzero diversion requirement is a pasted legacy value, not a formula result.",
                    asset.canonical_ditch_id, month,
                ))
        for month, is_formula in enumerate(asset.reservoir_net_evap_formula_months, 1):
            if not is_formula:
                issues.append(QAIssue(
                    "warning", "manual_reservoir_net_evap_override",
                    "Reservoir net evaporation is a legacy constant rather than the standard formula.",
                    asset.canonical_ditch_id, month,
                ))
    return tuple(issues)


def _find_report_ditch_column(sheet: object, report_ditch_name: str) -> int:
    for column in range(1, sheet.max_column + 1):
        if sheet.cell(6, column).value == report_ditch_name:
            return column
    raise ValueError(f"{sheet.title}: report ditch block {report_ditch_name!r} was not found in row 6.")


def _number(value: object) -> float:
    if not isinstance(value, (int, float)):
        if isinstance(value, str):
            try:
                return float(value)
            except ValueError:
                pass
        raise ValueError(f"Expected a numeric legacy report value, got {value!r}")
    return float(value)

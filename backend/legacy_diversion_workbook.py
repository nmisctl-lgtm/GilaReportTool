"""Read the 2024 daily-flow workbook only for baseline validation.

This adapter is intentionally narrow.  Future input formats should produce
``DailyDiversionRecord`` directly rather than forcing a report-year workbook
layout onto new source data.
"""

from __future__ import annotations

from datetime import date
from pathlib import Path

from .diversion_ingest import DailyDiversionRecord


# The legacy ``flow`` sheet explicitly labels daily records in rows 34--398.
# It also contains formatting below that range, which openpyxl may expose as
# used rows in read-only mode.
DAILY_FIRST_ROW = 34
DAILY_LAST_ROW = 398


def read_2024_flow_workbook(path: str | Path) -> tuple[DailyDiversionRecord, ...]:
    """Read the ``flow`` sheet's daily mean-CFS table into long-format records."""

    try:
        import openpyxl
    except ImportError as error:  # pragma: no cover - package is declared project dependency
        raise RuntimeError("Reading legacy Excel flow data requires the project's openpyxl dependency.") from error

    workbook = openpyxl.load_workbook(Path(path), data_only=True, read_only=True)
    sheet = workbook["flow"]
    headers = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
    names = {
        column: headers[column - 1]
        for column in range(2, len(headers) + 1)
        if headers[column - 1]
    }
    records: list[DailyDiversionRecord] = []
    for values in sheet.iter_rows(
        min_row=DAILY_FIRST_ROW,
        max_row=DAILY_LAST_ROW,
        max_col=len(headers),
        values_only=True,
    ):
        observed_on = values[0]
        if not isinstance(observed_on, date):
            continue
        if hasattr(observed_on, "date"):
            observed_on = observed_on.date()
        for column, ditch_name in names.items():
            value = values[column - 1]
            if value is not None and not isinstance(value, (int, float)):
                raise ValueError(f"{ditch_name} {observed_on}: expected numeric CFS, got {value!r}")
            records.append(DailyDiversionRecord(str(ditch_name), observed_on, value))
    return tuple(records)

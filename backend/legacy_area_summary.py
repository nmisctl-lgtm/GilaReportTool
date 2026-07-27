"""Read standard 2024 area-summary inputs from the legacy workbook.

This reader deliberately exposes the color-coded/manual yellow-box values as
data.  Future CropStats logic will replace this baseline adapter; calculation
code never needs to refer to Excel cells.
"""

from __future__ import annotations

from pathlib import Path

from .area_consumptive_use import (
    AcreageClass,
    CRPMeasuredUse,
    GenericAreaCUInput,
    MeteredSurfaceSupply,
    SpecialAreaCUInput,
)
from .legacy_report_assets import AREA_SHEET_NAMES


# Redrock and San Simon use additional CRP/full-CIR columns and are handled by
# a separate policy model.  The remaining seven sheets share this exact layout.
STANDARD_AREA_NAMES = (
    "LUNA", "APACHE-ARAGON", "RESERVE", "GLENWOOD", "UPPER GILA", "CLIFF-GILA", "VIRDEN VALLEY",
)


def read_2024_standard_area_inputs(path: str | Path) -> tuple[GenericAreaCUInput, ...]:
    """Extract the seven ordinary area layouts without using formula addresses downstream."""

    try:
        import openpyxl
    except ImportError as error:  # pragma: no cover - project dependency
        raise RuntimeError("Reading legacy area summaries requires openpyxl.") from error

    workbook = openpyxl.load_workbook(Path(path), data_only=True, read_only=False)
    results: list[GenericAreaCUInput] = []
    for area_name in STANDARD_AREA_NAMES:
        sheet = workbook[AREA_SHEET_NAMES[area_name]]
        crop_row, reservoir_row, total_row = _summary_rows(sheet)
        weather_total_row = next(
            row for row in range(1, crop_row)
            if sheet.cell(row, 1).value == "TOTALS"
        )
        results.append(GenericAreaCUInput(
            area_name=area_name,
            annual_cir_ft=_number(sheet.cell(crop_row, 3).value),
            annual_reservoir_net_evap_ft=(
                _number(sheet.cell(weather_total_row, 3).value)
                - _number(sheet.cell(weather_total_row, 4).value)
            ),
            metered_surface=MeteredSurfaceSupply(
                acreage=AcreageClass(
                    crop_acres=_number(sheet.cell(crop_row, 6).value),
                    reservoir_acres=_number(sheet.cell(reservoir_row, 6).value),
                ),
                diversion_required_acft=_number(sheet.cell(total_row, 8).value),
                diversion_shortage_acft=_number(sheet.cell(total_row, 9).value),
            ),
            unmetered_surface=AcreageClass(
                crop_acres=_number(sheet.cell(crop_row, 13).value),
                reservoir_acres=_number(sheet.cell(reservoir_row, 13).value),
            ),
            groundwater=AcreageClass(
                crop_acres=_number(sheet.cell(crop_row, 16).value),
                reservoir_acres=_number(sheet.cell(reservoir_row, 16).value),
            ),
            preplant_acres=_number(sheet.cell(crop_row + 1, 13).value),
        ))
    return tuple(results)


def read_2024_special_area_inputs(path: str | Path) -> tuple[SpecialAreaCUInput, ...]:
    """Extract the Redrock and San Simon full-CIR/CRP layouts."""

    try:
        import openpyxl
    except ImportError as error:  # pragma: no cover - project dependency
        raise RuntimeError("Reading legacy area summaries requires openpyxl.") from error

    workbook = openpyxl.load_workbook(Path(path), data_only=True, read_only=False)
    redrock = workbook[AREA_SHEET_NAMES["REDROCK"]]
    sansimon = workbook[AREA_SHEET_NAMES["SAN SIMON"]]
    return (
        _read_redrock(redrock),
        _read_san_simon(sansimon),
    )


def _read_redrock(sheet: object) -> SpecialAreaCUInput:
    weather_row = _weather_total_row(sheet)
    return SpecialAreaCUInput(
        area_name="REDROCK",
        annual_cir_ft=_number(sheet["C28"].value),
        annual_reservoir_net_evap_ft=_number(sheet.cell(weather_row, 3).value) - _number(sheet.cell(weather_row, 4).value),
        metered_surface=MeteredSurfaceSupply(
            acreage=AcreageClass(_number(sheet["F28"].value), _number(sheet["F30"].value)),
            diversion_required_acft=_number(sheet["H31"].value),
            diversion_shortage_acft=_number(sheet["I31"].value),
        ),
        unmetered_surface=AcreageClass(_number(sheet["L28"].value), _number(sheet["L30"].value)),
        groundwater=AcreageClass(_number(sheet["O28"].value), _number(sheet["O30"].value)),
        metered_full_cir=AcreageClass(_number(sheet["Q28"].value), _number(sheet["Q30"].value)),
        crp_measured_use=CRPMeasuredUse(
            acres=_number(sheet["S28"].value) + _number(sheet["S30"].value),
            measured_diversion_acft=_number(sheet["T28"].value) + _number(sheet["T30"].value),
        ),
    )


def _read_san_simon(sheet: object) -> SpecialAreaCUInput:
    weather_row = _weather_total_row(sheet)
    return SpecialAreaCUInput(
        area_name="SAN SIMON",
        annual_cir_ft=_number(sheet["C28"].value),
        annual_reservoir_net_evap_ft=_number(sheet.cell(weather_row, 3).value) - _number(sheet.cell(weather_row, 4).value),
        metered_surface=MeteredSurfaceSupply(AcreageClass(), 0.0, 0.0),
        metered_full_cir=AcreageClass(_number(sheet["L28"].value), _number(sheet["L30"].value)),
        groundwater=AcreageClass(_number(sheet["O28"].value), _number(sheet["O30"].value)),
        groundwater_cu_override_af=_number(sheet["P28"].value) + _number(sheet["P30"].value),
        crp_measured_use=CRPMeasuredUse(
            acres=_number(sheet["F28"].value) + _number(sheet["F30"].value),
            measured_diversion_acft=_zero(sheet["I28"].value) + _zero(sheet["I30"].value),
        ),
    )


def _summary_rows(sheet: object) -> tuple[int, int, int]:
    """Normal sheets put their Full Season row at 28 or 29."""

    crop_row = next(
        row for row in range(1, sheet.max_row + 1)
        if sheet.cell(row, 1).value == "Full Season" and sheet.cell(row, 3).value is not None
    )
    reservoir_row = next(
        row for row in range(crop_row + 1, min(crop_row + 4, sheet.max_row + 1))
        if sheet.cell(row, 1).value == "Reservoir"
    )
    total_row = next(
        row for row in range(reservoir_row + 1, min(reservoir_row + 3, sheet.max_row + 1))
        if sheet.cell(row, 1).value == "Total"
    )
    return crop_row, reservoir_row, total_row


def _weather_total_row(sheet: object) -> int:
    return next(row for row in range(1, sheet.max_row + 1) if sheet.cell(row, 1).value == "TOTALS")


def _number(value: object) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        return float(value)
    raise ValueError(f"Expected numeric legacy area value, got {value!r}")


def _zero(value: object) -> float:
    return 0.0 if value is None else _number(value)

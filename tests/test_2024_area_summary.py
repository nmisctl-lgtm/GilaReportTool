import unittest
from pathlib import Path

from backend.area_consumptive_use import calculate_generic_area_cu, calculate_special_area_cu
from backend.legacy_area_summary import read_2024_special_area_inputs, read_2024_standard_area_inputs
from backend.legacy_report_assets import AREA_SHEET_NAMES


WORKBOOK = Path(__file__).resolve().parents[1] / "OldMethod_Report/Spreadsheet/2024 Gila Report Data_WORKING.xlsx"


class StandardAreaSummaryTests(unittest.TestCase):
    def test_reproduces_all_seven_standard_2024_area_totals(self):
        inputs = read_2024_standard_area_inputs(WORKBOOK)
        self.assertEqual(len(inputs), 7)

        import openpyxl
        workbook = openpyxl.load_workbook(WORKBOOK, data_only=True, read_only=False)
        for area in inputs:
            result = calculate_generic_area_cu(area)
            sheet = workbook[AREA_SHEET_NAMES[area.area_name]]
            crop_row = next(
                row for row in range(1, sheet.max_row + 1)
                if sheet.cell(row, 1).value == "Full Season" and sheet.cell(row, 3).value is not None
            )
            total_row = crop_row + 3
            self.assertAlmostEqual(result.total_acres, sheet.cell(total_row, 19).value, places=7)
            self.assertAlmostEqual(result.full_supply_cu_af, sheet.cell(total_row, 20).value, places=7)
            self.assertAlmostEqual(result.shortage_to_cu_af, sheet.cell(total_row, 21).value, places=7)
            self.assertAlmostEqual(result.crop_and_pond_cu_af, sheet.cell(total_row, 22).value, places=7)
            self.assertAlmostEqual(result.incidental_losses_af, sheet.cell(total_row, 23).value, places=7)
            self.assertAlmostEqual(result.total_irrigated_cu_af, sheet.cell(total_row, 24).value, places=7)

    def test_reproduces_redrock_and_san_simon_special_class_totals(self):
        redrock, san_simon = read_2024_special_area_inputs(WORKBOOK)
        redrock_result = calculate_special_area_cu(redrock)
        san_simon_result = calculate_special_area_cu(san_simon)
        self.assertAlmostEqual(redrock_result.total_acres, 116.2, places=7)
        self.assertAlmostEqual(redrock_result.full_supply_cu_af, 337.17366666666663, places=7)
        self.assertAlmostEqual(redrock_result.shortage_to_cu_af, 130.6380934740303, places=7)
        self.assertAlmostEqual(redrock_result.crop_and_pond_cu_af, 206.53557319263633, places=7)
        self.assertAlmostEqual(san_simon_result.total_acres, 754.54, places=7)
        self.assertAlmostEqual(san_simon_result.full_supply_cu_af, 933.9736416666667, places=7)
        self.assertEqual(san_simon_result.shortage_to_cu_af, 0.0)
        self.assertAlmostEqual(san_simon_result.crop_and_pond_cu_af, 933.9736416666667, places=7)


if __name__ == "__main__":
    unittest.main()
